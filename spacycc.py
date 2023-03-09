from openpyxl import load_workbook
from openpyxl.styles import PatternFill
DATA_PATH = ".\\ossz.xlsx"
book = load_workbook(DATA_PATH,data_only=True)
sheet = book["Egybe"]
TYPES = ["FC0303","FC5203", "9DFC03"] #0 = Error 1 = Name error + 2
tanarnevek = ['Duret József','Ágoston György', 'Albert Sándor', 'Almási Tibor', 'Anderle Ádám', 'Bakos Ferenc', 'Balázs Mihály', 'Bálint Alajos', 'Bálint Sándor', 'Balogh Tibor', 'Banner János', 'Bánréti Zoltán', 'Baranyai Erzsébet', 'Baranyai Zoltán', 'Bárczi Géza', 'Barna Gábor', 'Baróti Dezső', 'Baróti Tiborné 1. Gaál Márta', 'Bartók György', 'Bassola Péter', 'Belényi Gyula', 'Bellon Tibor', 'Benedek Nándor', 'Bényiné 1. Farkas Mária', 'Bérezik Árpád', 'Bernáth Árpád', 'Berta Árpád', 'Birkás Géza', 'Boda István', 'Bognár Cecil Pál', 'Bollobás Enikő', 'Buday Árpád', 'Czachesz Erzsébet Cs.', 'Csapó Benő', 'Csefkó Gyula', 'Csejtei Dezső','Csengery János', 'Cseresnyési László', 'Csetri Lajos', 'Csukás István', 'Csúri Károly', 'Deér József', 'Deme László', 'Dézsi Lajos', 'Domokos Péter', 'Dornbach Mária', 'Duró Lajos', 'Eördögh István', 'Eperjessy Kálmán', 'Erdélyi Gyula', 'Erdélyi László', 'Erdélyi László Gyula', 'Erős Ferenc', 'Fabiny Tibor', 'Fábricz Károly', 'Farkas Mária', ' Bényiné', 'Fazekas Erzsébet', ' Gerő Ernőné', 'Fejér Ádám', 'Fejes Katalin', ' B.', 'Feleky Gábor', 'Felvinczi Takáts Zoltán', 'Fenyvesi István', 'Ferenczi Imre', 'Fodor István', 'Fogarasi Miklós', 'Fógel József', 'Fónagy Iván', 'Forgács Tamás', 'Förster Aurél', 'Gaál Endre', 'Gaál Márta', ' Baróti Tiborné', 'Galamb Sándor', 'Gausz András', 'Gazdapusztai Gyula', 'Gerő Ernőné 1. Fazekas Erzsébet', 'Göncz Lajos', 'Grezsa Ferenc', 'GundaBéla', 'Gyenge Zoltán', 'Gyimesi Sándor', 'Hahn István', 'Hajdú Péter', 'Hajnóczi Gábor', 'Halasy-Nagy József', 'Halász Előd', 'Halász Élődné 1. Szász Anna Mária', 'Hegyi András', 'Helembai Kornélia', 'Heller Ágnes', 'Hermann István Egyed', 'Herrmann Antal', 'Hoffmann Zsuzsanna', 'Horger Antal', 'Hornyánszky Gyula', 'Horváth István Károly', 'Horváth Károly', 'Huszti Dénes', 'Huszti József', 'Ilia Mihály', 'Imre Sándor', 'Imrényi Tibor', 'Ivanics Mária', 'Jakócs Dániel', 'Juhász Antal', 'Juhász Jenő József', 'Juhász József', 'Kanyó Zoltán', 'Kaposi Márton', 'Kardos (Pánd', 'Kari K.Lajos', 'Károly Sándor', 'Karsai László', 'Kecskeméti Ármin', 'Kékes Szabó Mihály', 'Kelemen János', 'Kenesei István', 'Kerényi Károly', 'Keserű Bálint', 'Király István', 'Kiss Lajos', 'Kissné 1. Nóvák Éva', 'Klemm Imre Antal', 'Kocziszky Éva Siflisné', 'Kocsis Mihály', 'Kocsondi András', 'Koltay-Kastner Jenő', 'Komlósy Ákos', 'Koncz János', 'Kontra Miklos', 'Kovács Ilona', 'Kozáky István', 'Kretzoi Sarolta', 'Kristó Gyula', 'Kukovecz Györgyné 1. Zentai Mária', 'Kürtösi Katalin', 'Lagzi Istvan', 'Lepahin Valerij', 'Madácsy László', 'Magyari Zoltánné 1. Techert Margit', 'Makk Ferenc', 'Mályusz Elemér', 'Marinovich Sarolta Resch Béláné', 'Marjanucz Lászlo', 'Márki Sándor', 'Maróti Egon', 'Martonyi Éva', 'Márvány János', 'Masát András', 'Meleczky Márta Judit', 'Mérei Gyula', 'Mester János', 'Mészáros Edit', 'Mészöly Gedeon', 'Mészöly Gedeon', 'Mezősi Károly', 'Mikola Tibor', 'Miskolczy István', 'Módi Mihály', 'Mokány Sándor', 'Nacsády József', 'Nagy Géza', 'Nagy József', 'Nagy László', 'Nagy László J.', 'Nagy Mária', ' Nagy Miklósné', 'Németh T. Enik', 'Nóvák Éva', ' Kissné', 'Nyíri Antal', 'Odorics Ferenc', 'Olajos Terézia', 'Olasz Sándor', 'Orosz Sándor', 'Oroszlán Zoltán', 'Ortutay Gyula', 'Ördögh Éva', 'Ötvös Péter', 'Pál József', 'Pálfy Miklós', 'Pándi Lajos', 'Párducz Mihály', 'Pável Ágoston', 'Penke Olga', ' Penke Botondné', 'Pete István', 'Péter László', 'Pordány László', 'Pósa Péter', 'Pukánszky Béla', 'Rácz Endre', 'Raffay Ernő', 'Resch Béláné 1.Marinovich Sarolta', 'Róna-Tas András', 'Roska Márton', 'Rózsa Éva', 'Rozsnyai Bálint', 'Rubinyi Mózes', 'Sajti Enikő', ' A.', 'Salyámosy Miklós', 'Siflisné 1. Kocziszky Éva', 'Szabó József (Magyar Nyelvészeti T', 'Szabó Tibor', 'Szádeczky-Kardoss Lajos', 'Szádeczky-Kardoss Samu', 'Szajbély Mihály', 'Szakáll Zsigmond', 'Szalamin Edit', 'Szalma Józsefné', ' Vihter Natalia', 'Szántó Imre', 'Szász Anna Mária', ' Halász Élődné', 'Szathmári István', 'Szauder József', 'Székely Lajos', 'Széles Klára', 'Szentiványi Róbert', 'Szerb Antal', 'Szigeti Lajos Sándor', 'Tar Ibolya', 'Techert Margit', ' Magyari Zoltánné', 'Tettamanti Béla', 'Timár Kálmán', 'Tóth Dezső', 'Tóth Imre H.', 'Tóth János', 'Tóth Sándor László', 'Trencsényi-Waldapfel Imre', 'Trogmayer Ottó', 'Varga Ilona', 'Veczkó József', 'Velcsov Mártonné 1. Tóth Katalin', 'Vértes O. József', 'Vidákovich Tibor', 'Vihter Natalia 1. Szalma Józsefné', 'Visy József', 'Vörös László', 'Zentai Mária', ' Kukovecz Györgyné', 'Zimonyi István', 'Zolnai Béla', 'Imre Sándor', 'Ladányi Gedeon', 'Szamosi  János', 'Szász Béla', 'Finaly Henrik', 'Hómao Ottó', 'Hómao Ottó', 'Szász Béla', 'Terner Adolf', 'Szabó Károly', 'Felméri Lajos', 'Szilasi Gergely', 'Hómao Ottó', 'Szamosi  János', 'Szász Béla', 'Hegedűs István', 'Dr. Meltzl Hugó', 'Hegedűs István', 'Szinnyei József', 'Schilling Lajos', 'Moldován Gergely', 'Széchy Károly', 'Pecz Vilmos', 'Szádeczky Lajos', 'Márki Sándor', 'Halász Ignác', 'Schneller István', 'Haraszti Gyula', 'Csngery János', 'Böhm Károly', 'Moldován Gergely', 'Vajda Gyula', 'Posta Béla', 'Schilling Lajos', 'Szádeczky Lajos', 'Márki Sándor', 'Schneller István', 'Haraszti Gyula', 'Posta Béla', 'Cholnoky Jenő', 'Zolnai Gyula', 'Dézsi Lajos', 'Schmidt Henrik', 'Zolnai Gyula', 'Erdélyi László', 'Hornyánszky Gyula', 'Finály Henrik Lajos']

def add_lines(line,rows):
  for i in range(2,rows+2):
    line.append(sheet[f"D{i}"].value)
def darabolo(line,parts):
    for i in range(len(line)):
      last = -1
      part = []
      if (line[i] != None):
          for c in range(len(line[i])):
              if line[i][c] == "," or line[i][c] == "." or line[i][c] == ";":
                  index = 0
                  #while line[i][last+1:c][index] == " " or line[i][last+1:c][index] == "\xa0":
                  #    index+=1
                  #print (index, last,i,c,line[i][last+1+index:c], last+1+index, c)
                  while True:
                    if (last+1+index != c):
                      if line[i][last+1:c][index] == " " or line[i][last+1:c][index] == "\xa0":
                        index += 1
                      else:
                        break
                    else:
                      break
                  #print(line[i][last+1+index:c])
                  if (last+1+index != c):
                    part.append(line[i][last+1+index:c])
                  last = c
      if (len(part) > 0):
          parts.append(part)
      else:
        parts.append(["NO DATA"])
    #print(parts)
def process(parts,napok,cimek,idok,rows):
  for d in range((len(parts))):
      #print(parts[d])
      index = 0
      van = False
      while True:
        for i in range(len(napok)):
          if (len(parts[d]) > index+1):
            if str.lower(parts[d][1+index]).rfind(napok[i]) > -1:
              van = True
              break
          else:
            van = True
            break
        if not(van):
          index += 1
        else:
          break
      szov = ""
      if(len(parts[d][0]) > 6):
        for i in range(0,index+1):
            if szov != "":
              szov = szov +", "+ parts[d][i]
            else:
              szov = szov + parts[d][i]
      else:
        for i in range(1,index+1):
            if szov != "":
              szov = szov +", "+ parts[d][i]
            else:
              szov = szov + parts[d][i]
      #cimek.append(parts[d][1])
      cimek.append(szov)
      #Név kezdete:Ugyanazon tanár.
      id = 0
      for i in range(index+1,len(parts[d])):
        if (parts[d][i] != None and sheet[f"G{d+2}"].value != None):
          if ((sheet[f"G{d+2}"].value).rfind(".") <= -1):
            if (str.lower(parts[d][i]).rfind("dr") > -1) or (str.lower(parts[d][i]).rfind(str.lower(sheet[f"G{d+2}"].value)) > -1) or (str.lower(parts[d][i]).rfind(str.lower("ugyana")) > -1) or (str.lower(parts[d][i]).rfind(str.lower("uyy")) > -1) or (str.lower(parts[d][i]).rfind(str.lower("ugyanom")) > -1):
            #if (str.lower(parts[d][i]).rfind(str.lower(sheet[f"G{d+2}"].value)) > -1) or (str.lower(parts[d][i]).rfind(str.lower("ugyan")) > -1):
              #print(i, parts[d][i])
              id = i
          else:
            shorter = sheet[f"G{d+2}"].value[:(sheet[f"G{d+2}"].value).rfind(".")]
            if (str.lower(parts[d][i]).rfind("dr") > -1) or (str.lower(parts[d][i]).rfind(str.lower(shorter)) > -1) or (str.lower(parts[d][i]).rfind(str.lower("ugyana")) > -1) or (str.lower(parts[d][i]).rfind(str.lower("uyy")) > -1) or (str.lower(parts[d][i]).rfind(str.lower("ugyanom")) > -1):
            #if (str.lower(parts[d][i]).rfind(str.lower(sheet[f"G{d+2}"].value)) > -1) or (str.lower(parts[d][i]).rfind(str.lower("ugyan")) > -1):
              #print(i, parts[d][i])
              id = i
      #print idő
      ido_str = ""
      if id > 0:
        sheet[f"G{d+2}"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
        for ido in range(index+1,id):
          if (ido_str != "" and parts[d][ido-1][len(parts[d][ido-1])-1] != "d"):
            ido_str = ido_str +", "+parts[d][ido]
          else:
            ido_str = ido_str + parts[d][ido]
          #print(parts[d][ido])
      else:
        if (sheet[f"G{d+2}"].value != None):
          ido_str = "NO DATA"
          sheet[f"F{d+2}"].fill = PatternFill(start_color=TYPES[1], end_color=TYPES[1], fill_type = "solid")
          sheet[f"G{d+2}"].fill = PatternFill(start_color=TYPES[2], end_color=TYPES[2], fill_type = "solid")
        else:
          ido_str = " "
      idok.append(ido_str)
def save(cimek,idok):
  for i in range(2,len(idok)+2):
    sheet[f"F{i}"].value = idok[i-2]
    if (sheet[f"F{i}"].value == "NO DATA"):
      sheet[f"F{i}"].fill = PatternFill(start_color=TYPES[0], end_color=TYPES[0], fill_type = "solid")
    else:
      sheet[f"F{i}"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
  for i in range(2,len(cimek)+2):
    sheet[f"E{i}"].value = cimek[i-2]
  book.save(DATA_PATH)
def name_check(rows):
    for i in range(2,rows+2):
        benne = False
        for name in tanarnevek:
            if (sheet[f"G{i}"].value == name):
                benne = True
        if (benne == False):
            sheet[f"H{i}"].value = "X"
            first = ""
            for c in range(len(sheet[f"H{i}"].value)):
                if sheet[f"H{i}"].value[c] == " ":
                    first = sheet[f"H{i}"].value[:c]
            for name in tanarnevek:
                if (name.rfind(first) > -1):
                    sheet[f"H{i}"].value = name
                    break
    book.save(DATA_PATH)
def calc(rows):
    no_data = 0
    for i in range(2, rows+2):
        if (sheet[f"F{i}"].value == "NO DATA"):
            no_data += 1
    print(f"Errors = {no_data}")
    print(f"{rows / no_data ,.00}% ({rows} / {no_data})")
def main(rows):
  line = []
  napok = ["hétfő", "kedd", "szerd",
           "csütörtök", "péntek", "szombat",
           "vasárnap", "hetenként", "hét",
           "óra", "órán", "heti", "het",
           "(folytatólag)", "mindennap",
           "minden", "délelőtt",
           "meghatározandó"]
  parts = []
  add_lines(line,rows)
  #print(line)
  #for l in range(len(line)):
    #print(line[l])
  darabolo(line,parts)
  process(parts,napok,cimek,idok,rows)
  save(cimek,idok)
  #name_check(rows)
  calc(rows)
  print("DONE")

cimek = []
idok = []
rows = 0
end = False
for row in sheet["D2":"D100000"]:
    for cell in row:
        if (cell.value == "END"):
            end = True
            break
        rows += 1
    if (end):
        break
main(rows)
#print(parts)
#print("Címek:",cimek)
#print("Idők:",idok)

# for l in range(len(parts)):
#     for tagok in parts[l]:
#         mostani_tagok = nlp(tagok)
#         for i in mostani_tagok.ents:
#             print(i,i.label_)
