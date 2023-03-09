from openpyxl.styles import PatternFill
from openpyxl import load_workbook
DATA_PATH = ".\\prototype.xlsx"
book = load_workbook(DATA_PATH,data_only=True)
sheet = book["Munka1"]
types = ["FC0303","FC5203", "9DFC03"] #0 = Error 1 = Name error + 2 
sheet["G3"].fill = PatternFill(start_color=types[0], end_color=types[0], fill_type = "solid")
book.save(DATA_PATH)
