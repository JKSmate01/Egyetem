import spacy
import huspacy
from openpyxl import load_workbook
#huspacy.download
nlp = spacy.load("hu_core_news_lg")
#book = load_workbook('.\egyetem\ossz.xlsx,data_only=True)
#sheet = book["Egybe"]
DATA_PATH = ".\\ossz.xlsx"
book = load_workbook(DATA_PATH,data_only=True)
sheet = book["Egybe"]

def add_lines(line,rows):
  for i in range(2,rows+2):
    line.append(sheet[f"D{i}"].value)
def darabolo_JM(line,parts):
    for i in range(len(line)):
      last = -1
      part = []
      if (line[i] != None):
          for c in range(len(line[i])):
              if line[i][c] == "," or line[i][c] == "." or line[i][c] == ";":
                  index = 0
                  #while line[i][last+1:c][index] == " " or line[i][last+1:c][index] == "\xa0":
                  #    index+=1
                  #print (index, last,i,c,line[i][last+1+index:c], last+1+index, c)
                  while True:
                    if (last+1+index != c):
                      if line[i][last+1:c][index] == " " or line[i][last+1:c][index] == "\xa0":
                        index += 1
                      else:
                        break
                    else:
                      break
                  #print(line[i][last+1+index:c])
                  if (last+1+index != c):
                    part.append(line[i][last+1+index:c])
                  last = c
      if (len(part) > 0):
          parts.append(part)
      else:
        parts.append(["NO DATA"])
    #print(parts)
def process(parts,napok,cimek,idok,rows):
  for d in range((len(parts))):
      #print(parts[d])
      index = 0
      van = False
      while True:
        for i in range(len(napok)):
          if (len(parts[d]) > index+1):
            if str.lower(parts[d][1+index]).rfind(napok[i]) > -1:
              van = True
              break
          else:
            van = True
            break
        if not(van):
          index += 1
        else:
          break
      szov = ""
      for i in range(1,index+1):
          if szov != "":
            szov = szov +", "+ parts[d][i]
          else:
            szov = szov + parts[d][i]
      #cimek.append(parts[d][1])
      cimek.append(szov)
      #Név kezdete:Ugyanazon tanár.
      id = 0
      for i in range(index+1,len(parts[d])):
        if (parts[d][i] != None and sheet[f"G{d+2}"].value != None):
          if (str.lower(parts[d][i]).rfind("dr") > -1) or (str.lower(parts[d][i]).rfind(str.lower(sheet[f"G{d+2}"].value)) > -1) or (str.lower(parts[d][i]).rfind(str.lower("ugyana")) > -1):
          #if (str.lower(parts[d][i]).rfind(str.lower(sheet[f"G{d+2}"].value)) > -1) or (str.lower(parts[d][i]).rfind(str.lower("ugyan")) > -1):
            #print(i, parts[d][i])
            id = i
      #print idő
      ido_str = ""
      if id > 0:
        for ido in range(index+1,id):
          if (ido_str != "" and parts[d][ido-1][len(parts[d][ido-1])-1] != "d"):
            ido_str = ido_str +", "+parts[d][ido]
          else:
            ido_str = ido_str + parts[d][ido]
          #print(parts[d][ido])
      else:
        if (sheet[f"G{d+2}"].value != None):
          ido_str = "NO DATA"
        else:
          ido_str = " "
      idok.append(ido_str)
def save(cimek,idok):
  for i in range(2,len(idok)+2):
    sheet[f"F{i}"].value = idok[i-2]
  for i in range(2,len(cimek)+2):
    sheet[f"E{i}"].value = cimek[i-2]
  book.save(DATA_PATH)
def main(rows):
  line = []
  napok = ["hétfő","kedd","szerd","csütörtök","péntek","szombat","vasárnap","hetenként","hét", "óra","heti","het","(folytatólag)","mindennap","minden", "délelőtt"]
  parts = []
  add_lines(line,rows)
  print(line)
  #for l in range(len(line)):
    #print(line[l])
  darabolo_JM(line,parts)
  process(parts,napok,cimek,idok,rows)
  save(cimek,idok)
cimek = []
idok = []
rows = 0
for row in sheet["D2":"D100000"]:
    for cell in row:
        rows += 1
main(rows)
#print(parts)
print("Címek:",cimek)
print("Idők:",idok)

# for l in range(len(parts)):
#     for tagok in parts[l]:
#         mostani_tagok = nlp(tagok)
#         for i in mostani_tagok.ents:
#             print(i,i.label_)
